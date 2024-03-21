import openpyxl
from .log_uitl import *


class ExcelHelper:
    def __init__(self):
        self.workbook = None
        self.logger = init_logging

    def load_workbook(self, filename):
        """
        加载工作簿

        :param filename: 工作簿文件路径
        """
        self.logger.info(f'从 {filename} 加载工作簿')
        self.workbook = openpyxl.load_workbook(filename)

    def get_sheet(self, sheet_name=None):
        """
        获取工作表

        :param sheet_name: 工作表名称
        :return: 工作表对象
        """
        if sheet_name is None:
            sheet_name = self.workbook.active.title
        self.logger.info(f'获取工作表 {sheet_name}')
        return self.workbook[sheet_name]

    def get_cell_value(self, sheet_name, row, column):
        """
        获取单元格值

        :param sheet_name: 工作表名称
        :param row: 行号
        :param column: 列号
        :return: 单元格值
        """
        sheet = self.get_sheet(sheet_name)
        self.logger.info(f'从工作表 {sheet_name} 的第 {row} 行、第 {column} 列获取单元格值')
        return sheet.cell(row, column).value

    def set_cell_value(self, sheet_name, row, column, value):
        """
        设置单元格值

        :param sheet_name: 工作表名称
        :param row: 行号
        :param column: 列号
        :param value: 要设置的值
        """
        sheet = self.get_sheet(sheet_name)
        self.logger.info(f'在工作表 {sheet_name} 的第 {row} 行、第 {column} 列设置单元格值为 {value}')
        sheet.cell(row, column).value = value

    def save(self):
        """
        保存工作簿
        """
        self.logger.info('保存工作簿')
        self.workbook.save()

    def get_sheet_names(self):
        """
        获取所有工作表名称

        :return: 工作表名称列表
        """
        self.logger.info('获取工作表名称')
        return self.workbook.sheetnames

    def create_sheet(self, sheet_name):
        """
        创建工作表

        :param sheet_name: 工作表名称
        """
        self.logger.info(f'创建工作表 {sheet_name}')
        self.workbook.create_sheet(sheet_name)

    def delete_sheet(self, sheet_name):
        """
        删除工作表

        :param sheet_name: 工作表名称
        """
        self.logger.info(f'删除工作表 {sheet_name}')
        self.workbook.remove(self.workbook[sheet_name])
